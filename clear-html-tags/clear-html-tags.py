from openpyxl import load_workbook
from bs4 import BeautifulSoup


# Function to remove tags
def remove_tags(html):
  
    # parse html content
    soup = BeautifulSoup(html, "html.parser")
  
    for data in soup(['style', 'script']):
        # Remove tags
        data.decompose()
  
    # return data by retrieving the tag content
    return ' '.join(soup.stripped_strings)


if __name__ == '__main__':

    # Para evitar alterar sempre o código fonte, posicione a coluna com o 
    # conteúdo HTML na primeira coluna da planilha.
    # Deixa a segunda coluna da planilha em branco, pois será onde o texto
    # corrigido será gravado.

    wb = load_workbook(filename='planilha.xlsx')
    for ws in wb.worksheets:
        for index, row in enumerate(ws.rows, start=1):
            #print(row[0].value)
            #print( remove_tags(row[0].value) ) 
            #print(index, ws.cell(row=index, column=2).value)
            x1 = remove_tags(row[0].value)
            ws.cell(row=index, column=2).value = x1

    # Save a new file
    wb.save(filename='planilha_fixed.xlsx')
    wb.close()